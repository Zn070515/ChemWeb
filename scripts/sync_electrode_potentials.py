import json
import re
import sys
from pathlib import Path

import openpyxl

EXCEL_PATH = Path(r"C:\\Users\\16275\\Desktop\\最新最全最实用电极电势表.xlsx")
DATA_PATH = Path(r"C:\\Users\\16275\\Desktop\\ChemWeb\\public\\elements-data.json")


def normalize_reaction(value: str) -> str:
    text = str(value)
    text = text.replace("＋", "+").replace("－", "-").replace("＝", "=")
    text = text.replace("﹢", "+").replace("﹣", "-")
    text = re.sub(r"\s+", " ", text).strip()
    text = text.replace(" +", " + ").replace("+ ", " + ")
    text = text.replace(" =", " = ").replace("= ", " = ")
    text = re.sub(r"\s+", " ", text).strip()
    text = re.sub(r"\s+\(", "(", text)
    text = re.sub(r"([A-Za-z0-9\)\]])\s+([+-])", r"\1\2", text)
    return text


def normalize_charge(charge: str) -> str:
    if charge in {"+", "-"}:
        return charge
    match = re.fullmatch(r"([+-])(\d+)", charge)
    if match:
        sign, digits = match.groups()
        return f"{digits}{sign}"
    match = re.fullmatch(r"(\d+)([+-])", charge)
    if match:
        digits, sign = match.groups()
        return f"{digits}{sign}"
    return charge


def split_charge(text: str) -> tuple[str, str | None]:
    match = re.fullmatch(r"(.*?)([+-]\d*)", text)
    if match:
        return match.group(1), match.group(2)

    match = re.fullmatch(r"(.*?)(\d+)([+-])", text)
    if match:
        base, digits, sign = match.groups()
        elements = re.findall(r"[A-Z][a-z]?", base)
        if len(elements) > 1:
            return f"{base}{digits}", sign
        return base, f"{digits}{sign}"
    return text, None


def format_formula(formula: str) -> str:
    match = re.match(r"^(\d+)(.*)$", formula)
    coeff = ""
    if match:
        coeff, formula = match.groups()

    formula = re.sub(r"([A-Z][a-z]?)(\d+)", r"\1<sub>\2</sub>", formula)
    formula = re.sub(r"(\))(\d+)", r"\1<sub>\2</sub>", formula)
    return f"{coeff}{formula}"


def format_species(token: str) -> str:
    token = token.strip()
    if not token:
        return token

    suffix = ""
    suffix_match = re.search(r"(\([^)]*\))$", token)
    if suffix_match:
        suffix = suffix_match.group(1)
        token = token[: -len(suffix)]

    base, charge = split_charge(token)
    base = base.strip()
    if base == "e":
        if charge:
            return f"e<sup>{normalize_charge(charge)}</sup>{suffix}"
        return f"e{suffix}"

    formatted = format_formula(base)
    if charge:
        formatted = f"{formatted}<sup>{normalize_charge(charge)}</sup>"
    return f"{formatted}{suffix}"


def format_reaction_html(reaction_text: str) -> str:
    text = reaction_text.strip()
    tokens: list[str] = []
    buffer: list[str] = []
    in_paren = 0

    def flush_buffer() -> None:
        if buffer:
            tokens.append("".join(buffer).strip())
            buffer.clear()

    length = len(text)
    for idx, char in enumerate(text):
        if char == "(":
            in_paren += 1
            buffer.append(char)
            continue
        if char == ")" and in_paren:
            in_paren -= 1
            buffer.append(char)
            continue
        if char in {"+", "="} and not in_paren:
            if char == "=":
                flush_buffer()
                tokens.append("=")
                continue

            next_non_space = None
            for j in range(idx + 1, length):
                if text[j] != " ":
                    next_non_space = text[j]
                    break
            prev_non_space = None
            for j in range(idx - 1, -1, -1):
                if text[j] != " ":
                    prev_non_space = text[j]
                    break

            if next_non_space in {None, "+", "="}:
                buffer.append(char)
                continue
            if next_non_space and (next_non_space.isalpha() or next_non_space.isdigit()):
                flush_buffer()
                tokens.append("+")
                continue
            if prev_non_space in {None, "(", "=", "+"}:
                buffer.append(char)
                continue
            flush_buffer()
            tokens.append("+")
            continue

        buffer.append(char)

    flush_buffer()

    formatted_parts: list[str] = []
    for token in tokens:
        if token in {"+", "="}:
            formatted_parts.append(f" {token} ")
        else:
            formatted_parts.append(format_species(token))

    return re.sub(r"\s+", " ", "".join(formatted_parts)).strip()


def normalize_value(value) -> str:
    if isinstance(value, (int, float)):
        return str(value)
    text = str(value).strip()
    text = text.replace("（", "(").replace("）", ")")
    text = re.sub(r"\(V\)", "", text, flags=re.IGNORECASE)
    text = text.replace("V", "").strip()
    return text


def parse_elements(text: str) -> list[str]:
    return re.findall(r"[A-Z][a-z]?", text)


def is_simple_element(text: str) -> str | None:
    clean = text.strip()
    match = re.fullmatch(r"([A-Z][a-z]?)(\d+)?([+-].*)?", clean)
    if not match:
        return None
    elements = parse_elements(clean)
    if len(elements) != 1:
        return None
    return elements[0]


def pick_symbol_from_couple(couple_text: str) -> str | None:
    if "/" not in couple_text:
        elements = parse_elements(couple_text)
        if not elements:
            return None
        return elements[0]

    left, right = couple_text.split("/", 1)
    right_symbol = is_simple_element(right)
    if right_symbol:
        return right_symbol
    left_symbol = is_simple_element(left)
    if left_symbol:
        return left_symbol

    left_elements = parse_elements(left)
    if left_elements:
        for symbol in left_elements:
            if symbol != "H":
                return symbol
        return left_elements[0]

    right_elements = parse_elements(right)
    if right_elements:
        for symbol in right_elements:
            if symbol != "H":
                return symbol
        return right_elements[0]
    return None


def build_mapping() -> dict[str, list[str]]:
    workbook = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    worksheet = workbook.active
    rows = list(worksheet.iter_rows(values_only=True))
    header_index = 0
    for idx, row in enumerate(rows):
        if row and "电对符号" in row:
            header_index = idx
            break
    data_rows = [row for row in rows[header_index + 1 :] if any(row)]

    mapping: dict[str, list[str]] = {}
    for couple, reaction, potential, _note in data_rows:
        if not couple or not reaction or potential is None:
            continue
        couple_text = str(couple).strip()
        symbol = pick_symbol_from_couple(couple_text)
        if not symbol:
            continue
        reaction_text = normalize_reaction(reaction)
        reaction_text = format_reaction_html(reaction_text)
        potential_text = normalize_value(potential)
        if not reaction_text or potential_text is None:
            continue
        entry = f"{reaction_text}（E<sup>θ</sup> = {potential_text} V）"
        mapping.setdefault(symbol, []).append(entry)
    return mapping


def update_data_file(mapping: dict[str, list[str]]) -> None:
    lines = DATA_PATH.read_text(encoding="utf-8").splitlines(keepends=True)
    current_symbol = None
    updated_lines = []

    for line in lines:
        symbol_match = re.search(r'"symbol"\s*:\s*"([A-Za-z]{1,2})"', line)
        if symbol_match:
            current_symbol = symbol_match.group(1)

        if '"standardElectrodePotential"' in line and current_symbol:
            entries = mapping.get(current_symbol)
            if entries:
                value = "<br>".join(entries)
                indent = line.split('"standardElectrodePotential"')[0]
                json_value = json.dumps(value, ensure_ascii=False)
                line = f"{indent}\"standardElectrodePotential\":  {json_value}\n"
        updated_lines.append(line)

    DATA_PATH.write_text("".join(updated_lines), encoding="utf-8")


def main() -> None:
    if not EXCEL_PATH.exists():
        raise SystemExit(f"Excel file not found: {EXCEL_PATH}")
    mapping = build_mapping()
    if len(sys.argv) > 1 and sys.argv[1] == "debug-h":
        workbook = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
        worksheet = workbook.active
        rows = list(worksheet.iter_rows(values_only=True))
        header_index = 0
        for idx, row in enumerate(rows):
            if row and "电对符号" in row:
                header_index = idx
                break
        data_rows = [row for row in rows[header_index + 1 :] if any(row)]
        for couple, reaction, potential, note in data_rows:
            if couple and str(couple).strip().startswith("H"):
                print(couple, reaction, potential, note)
        return
    update_data_file(mapping)


if __name__ == "__main__":
    main()
